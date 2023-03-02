from selenium import webdriver
from bs4 import BeautifulSoup
import time
import csv
import re
import datetime
import random
import pandas as pd
import openpyxl

#utworzenie obiektu przeglądarki, wraz z dzisiejszą datą, która będzie użyta jako nazwa zakładki w pliku excel
browser = webdriver.Chrome()
now = datetime.datetime.now()
current_date = now.strftime("%Y-%m-%d")


#otwarcie strony z ogłoszeniami w pętli ( od pierwszej do piątej strony )
df = pd.DataFrame(columns=['tytul_ogloszenia', 'Lokalizacja', 'Cena', 'link_ogloszenia', 'Data','Powierzchnia','liczba_pokoi','Pietro','Rok budowy'])
book = openpyxl.load_workbook('gratka_kielce.xlsx')
if current_date not in book.sheetnames:
    book.create_sheet(current_date)
else:
    print('Ta data już została pobrana do skoroszytu!')
    exit()

for i in range(1,3):
#otwarcie strony z ogłoszeniami
    browser.get(f'https://gratka.pl/nieruchomosci/mieszkania/kielce?page={i}&sort=newest')
    time.sleep(random.randint(30,40))

#pobranie surowego kodu html
    html = browser.page_source

#przetworzenie kodu html przy pomocy BeautifulSoup
    soup = BeautifulSoup(html, 'html.parser')

#pobranie wszystkich ogłoszeń znajdujących się na stronie
    results = soup.find_all('div', {'class': 'listing__teaserWrapper'})

#otwarcie pliku excel


    for result in results:
            title = result.find('a', {'class': 'teaserLink'})['title']

            area = result.find('span', {'class': 'teaserUnified__location'}).text
            area = re.sub(r'\s', '', area)
            area = re.sub(r',', ', ', area)

            price = result.find('p', {'class': 'teaserUnified__price'}).contents[0]
            price = price.replace(' ','').replace('\n','')
            if 'Zapytajocenę' in price:
                price = 'NULL'
            else:
                price
        
            
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument('--new-tab')
            browser = webdriver.Chrome(chrome_options=chrome_options)

            link = result.find('a', {'class': 'teaserLink'})
            browser.execute_script("window.open('{}');".format(link['href']))
            print(link)
            
            time.sleep(random.randint(6,8))

            handles = browser.window_handles
            # Przełączenie się na nowo otwartą kartę
            browser.switch_to.window(browser.window_handles[-1])

            # Pobranie kodu HTML z nowej karty
            html = browser.page_source
            
            # Przetworzenie kodu HTML przy pomocy BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')


             
            # Pobranie danych z nowej karty
            powierzchnia = soup.select_one("li span:contains('Powierzchnia w m2') + b").text
            powierzchnia = re.sub(r' m2', '', powierzchnia)
            powierzchnia = powierzchnia if powierzchnia else "NULL"

        
            liczba_pokoi = soup.select_one("li span:contains('Liczba pokoi') + b").text
            liczba_pokoi = liczba_pokoi if liczba_pokoi else "NULL"


            pietro = soup.select_one("li span:contains('Piętro') + b")
            if pietro is None:
                pietro = "NULL"
            elif 'parter' in pietro:
                pietro = '0'
            else:
                pietro = pietro.text


            rok_budowy = soup.select_one("li span:contains('Rok budowy') + b")
            rok_budowy = rok_budowy.text if rok_budowy else "NULL"
            
            data = [title, area, price, link['href'], current_date]
            data.append(powierzchnia or 'NULL')
            data.append(liczba_pokoi or 'NULL')
            data.append(pietro or 'NULL')
            data.append(rok_budowy or 'NULL')

    # Tworzenie DataFrame z pobranymi danymi i zapis do pliku excel
            
            
            writer = pd.ExcelWriter('gratka_kielce.xlsx', engine='openpyxl')
            writer.book = book
            df = df.append({'tytul_ogloszenia': title, 'Lokalizacja': area, 'Cena': price, 'link_ogloszenia': link['href'], 'Data': current_date,'Powierzchnia':powierzchnia,'liczba_pokoi':liczba_pokoi,'Pietro':pietro,'Rok budowy':rok_budowy}, ignore_index=True)
            df.to_excel(writer, sheet_name=current_date, index=False, header=True)
            writer.save()
browser.close()
 
    





        

"""

 writer.writerow([title, area, price, link['href']])

from selenium import webdriver
from bs4 import BeautifulSoup
import time
import csv

#utworzenie obiektu przeglądarki
browser = webdriver.Chrome()

#otwarcie strony z ogłoszeniami
browser.get('https://gratka.pl/nieruchomosci/mieszkania/kielce')


#pobranie surowego kodu html
html = browser.page_source

#przetworzenie kodu html przy pomocy BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')

#pobranie wszystkich ogłoszeń znajdujących się na stronie
results = soup.find_all('div', {'class': 'listing__teaserWrapper'})

#otwarcie pliku csv
with open('gratka_kielce.csv', mode='w', newline='') as csv_file:
    fieldnames = ['tytul_ogloszenia', 'Lokalizacja', 'Cena', 'link_ogloszenia']
    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)

    writer.writeheader()

    #przeglądanie ogłoszeń i zapisywanie informacji do pliku csv
for result in results:
    titles = soup.find_all('a', {'class': 'teaserLink'})
    area = soup.find_all('span', {'class': 'teaserUnified__location'})
    price = soup.find_all('p', {'class': 'teaserUnified__price'}) 
    link = soup.find_all('a', {'class': 'teaserLink'})
    
for pric in price:
        price_text = pric.split()[0] + pric.split()[1]
        print(price_text)

for title in titles:
    print(title['title'])


for are in area:
    print(are.text)
"""